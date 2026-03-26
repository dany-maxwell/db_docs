from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QTableWidget, QTableWidgetItem, 
    QPushButton, QComboBox, QDateEdit,
    QMessageBox, QFileDialog, QHBoxLayout
)
from PySide6.QtCore import QTimer
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment

from services.busqueda_service import busqueda_documentos_avanzada
from services.catalogo_service import catalogo_subtipos, catalogo_reporte
from ui.widgets.widgets import FormularioBusqueda
from constants import HEADER_LABELS_CONSULTAR, INDICES_DB_CONSULTAR, MSG_EXCEL_EXPORTADO, MSG_TITULO_RESUMEN

class WidgetConsultar(QWidget):
    def __init__(self):
        super().__init__()
        self.setup_ui()
        self.timer_busqueda = QTimer()
        self.timer_busqueda.setSingleShot(True)
        self.timer_busqueda.timeout.connect(self.buscar)
        self.conectar_eventos()
        self.buscar()

    def setup_ui(self):
        layout = QHBoxLayout(self)
        
        self.filtros = FormularioBusqueda()
        self.filtros.setFixedWidth(300)
        layout.addWidget(self.filtros)

        derecha = QVBoxLayout()
        self.tabla = QTableWidget(0, len(HEADER_LABELS_CONSULTAR))
        self.tabla.setHorizontalHeaderLabels(HEADER_LABELS_CONSULTAR)
        self.tabla.horizontalHeader().setStretchLastSection(True)
        self.tabla.setAlternatingRowColors(True)
        self.tabla.setEditTriggers(QTableWidget.NoEditTriggers)
        
        self.botoneslayout = QHBoxLayout()
        self.btn_excel = QPushButton("Exportar Tabla Actual")
        self.btn_reporte = QPushButton("Generar Reporte")
        self.botoneslayout.addWidget(self.btn_excel)
        self.botoneslayout.addWidget(self.btn_reporte)
        derecha.addWidget(self.tabla)
        derecha.addLayout(self.botoneslayout)
        layout.addLayout(derecha)

    def conectar_eventos(self):
        for cb in self.filtros.findChildren(QComboBox):
            cb.currentIndexChanged.connect(self.disparar_busqueda)

        for de in self.filtros.findChildren(QDateEdit):
            de.dateChanged.connect(self.disparar_busqueda)

        self.filtros.txt_codigo.textChanged.connect(self.disparar_busqueda)
        self.filtros.btn_limpiar.clicked.connect(self.filtros.limpiar)
        self.filtros.btn_limpiar.clicked.connect(self.buscar)
        self.filtros.combo_tipo.currentIndexChanged.connect(self.actualizar_subtipos)
        self.btn_excel.clicked.connect(self.exportar_excel)
        self.btn_reporte.clicked.connect(self.exportar_reporte)

    def disparar_busqueda(self):
        self.timer_busqueda.start(400)

    def actualizar_subtipos(self):
        id_tipo = self.filtros.combo_tipo.currentData()
        self.filtros.combo_sub.actualizar_items(catalogo_subtipos(id_tipo))

    def buscar(self):
        params = self.filtros.obtener_filtros()
        datos = busqueda_documentos_avanzada(**params)
        self.cargar_tabla(datos)

    def actualizar_proveedores(self):
        self.filtros.actualizar_combos()

    def actualizar_combos(self):
        self.filtros.actualizar_combos()
        self.buscar()

    def cargar_tabla(self, datos):
        self.tabla.setRowCount(0)
        self.tabla.setUpdatesEnabled(False)
        
        for fila_data in datos:
            row_idx = self.tabla.rowCount()
            self.tabla.insertRow(row_idx)
            
            for col_idx, db_idx in enumerate(INDICES_DB_CONSULTAR):
                valor = fila_data[db_idx]
                texto = str(valor) if valor is not None else ""
                self.tabla.setItem(row_idx, col_idx, QTableWidgetItem(texto))
                
        self.tabla.setUpdatesEnabled(True)

    def exportar_excel(self):
        ruta, _ = QFileDialog.getSaveFileName(
            self, "Guardar Excel", "consulta_tramites.xlsx", "Excel (*.xlsx)"
        )
        if not ruta:
            return

        wb = Workbook()
        id_memo = self.filtros.combo_memo.currentData()

        if id_memo:
            self._crear_hoja_resumen(wb)
            self._crear_hoja_documentos(wb, "Documentos")
        else:
            self._crear_hoja_documentos(wb, "Listado")

        wb.save(ruta)
        QMessageBox.information(self, "Exportado", MSG_EXCEL_EXPORTADO)

    def _crear_hoja_resumen(self, wb):
        ws = wb.active
        ws.title = "Resumen"
        ws["A1"] = MSG_TITULO_RESUMEN
        ws["A1"].font = Font(bold=True, size=14)

        if self.tabla.rowCount() == 0:
            ws["A3"] = "No hay datos cargados"
            return

        datos = [
            ("N° Trámite", self.tabla.item(0, 0).text()),
            ("Proveedor", self.tabla.item(0, 1).text()),
            ("Unidad", self.tabla.item(0, 2).text()),
            ("Fecha inicio", self.tabla.item(0, 3).text()),
        ]

        for i, (titulo, valor) in enumerate(datos, start=3):
            ws[f"A{i}"] = titulo
            ws[f"B{i}"] = valor
            ws[f"A{i}"].font = Font(bold=True)

        ws["A8"] = "Total documentos:"
        ws["B8"] = self.tabla.rowCount()

    def _crear_hoja_documentos(self, wb, nombre):
        ws = wb.create_sheet(nombre)
        ws.append(HEADER_LABELS_CONSULTAR)

        for cells in ws.iter_rows(min_row=1, max_row=1):
            for cell in cells:
                cell.font = Font(bold=True)

        for row in range(self.tabla.rowCount()):
            ws.append([
                self.tabla.item(row, col).text() 
                for col in range(self.tabla.columnCount())
            ])

        for column_cells in ws.columns:
            max_len = max(
                (len(str(cell.value or "")) for cell in column_cells),
                default=1
            )
            ws.column_dimensions[column_cells[0].column_letter].width = max_len + 2

    def exportar_reporte(self):
        ruta, _ = QFileDialog.getSaveFileName(
            self, "Guardar Excel", "reporte_tramites.xlsx", "Excel (*.xlsx)"
        )
        if not ruta:
            return

        wb = Workbook()

        datos = catalogo_reporte()
        ws = wb.active
        ws.title = "Reporte Trámites"

        headers = datos['columnas']

        ws.append(headers)

        for row in datos['rows']:
            ws.append(row)

        thin = Side(style='thin')
        borde = Border(
            left=thin,
            right=thin,
            top=thin,
            bottom=thin
        )

        for row in ws.iter_rows():
            for cell in row:
                cell.border = borde

        for cell in ws[1]:
            cell.font = Font(bold=True, size=12, color="FFFFFF")
            cell.fill = PatternFill(start_color="444444", end_color="444444", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        fill_resuelto = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        fill_actuacionPrevia = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        fill_instruccion = PatternFill(start_color="66FFFF", end_color="66FFFF", fill_type="solid")
        fill_resolucion = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")

        head = [cell.value for cell in ws[1]]
        col_estado = head.index("ESTADO")

        for row in ws.iter_rows(min_row=2):
            estado = row[col_estado].value

            if estado == "RESUELTO":
                for cell in row:
                    cell.fill = fill_resuelto
            if estado == "EN ACTUACIÓN PREVIA":
                for cell in row:
                    cell.fill = fill_actuacionPrevia
            if estado == "EN INSTRUCCIÓN":
                for cell in row:
                    cell.fill = fill_instruccion
            if estado == "EN RESOLUCIÓN":
                for cell in row:
                    cell.fill = fill_resolucion


        wb.save(ruta)
        QMessageBox.information(self, "Exportado", MSG_EXCEL_EXPORTADO)