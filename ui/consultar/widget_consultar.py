from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QTableWidget, QTableWidgetItem, 
    QPushButton, QComboBox, QDateEdit,
    QMessageBox, QFileDialog, QHBoxLayout
)
from PySide6.QtCore import QTimer
from openpyxl import Workbook
from openpyxl.styles import Font

from services.busqueda_service import busqueda_documentos_avanzada
from services.catalogo_service import catalogo_subtipos
from ui.widgets import FormularioBusqueda
from constants import HEADER_LABELS_CONSULTAR, INDICES_DB_CONSULTAR, MSG_EXCEL_EXPORTADO, MSG_TITULO_RESUMEN

class WidgetConsultar(QWidget):
    def __init__(self):
        super().__init__()
        self.setup_ui()
        self.timer_busqueda = QTimer()
        self.timer_busqueda.setSingleShot(True)
        self.timer_busqueda.timeout.connect(self.buscar)
        self.conectar_eventos()
        self.buscar()

    def setup_ui(self):
        layout = QHBoxLayout(self)
        
        self.filtros = FormularioBusqueda()
        self.filtros.setFixedWidth(300)
        layout.addWidget(self.filtros)

        derecha = QVBoxLayout()
        self.tabla = QTableWidget(0, len(HEADER_LABELS_CONSULTAR))
        self.tabla.setHorizontalHeaderLabels(HEADER_LABELS_CONSULTAR)
        self.tabla.horizontalHeader().setStretchLastSection(True)
        self.tabla.setAlternatingRowColors(True)
        
        self.btn_excel = QPushButton("Exportar a Excel")
        derecha.addWidget(self.tabla)
        derecha.addWidget(self.btn_excel)
        layout.addLayout(derecha)

    def conectar_eventos(self):
        for cb in self.filtros.findChildren(QComboBox):
            cb.currentIndexChanged.connect(self.disparar_busqueda)

        for de in self.filtros.findChildren(QDateEdit):
            de.dateChanged.connect(self.disparar_busqueda)

        self.filtros.txt_codigo.textChanged.connect(self.disparar_busqueda)
        self.filtros.btn_limpiar.clicked.connect(self.filtros.limpiar)
        self.filtros.btn_limpiar.clicked.connect(self.buscar)
        self.filtros.combo_tipo.currentIndexChanged.connect(self.actualizar_subtipos)
        self.btn_excel.clicked.connect(self.exportar_excel)

    def disparar_busqueda(self):
        self.timer_busqueda.start(400)

    def actualizar_subtipos(self):
        id_tipo = self.filtros.combo_tipo.currentData()
        self.filtros.combo_sub.actualizar_items(catalogo_subtipos(id_tipo))

    def buscar(self):
        params = self.filtros.obtener_filtros()
        datos = busqueda_documentos_avanzada(**params)
        self.cargar_tabla(datos)

    def actualizar_combos(self):
        self.filtros.actualizar_combos()
        self.buscar()

            
    def cargar_tabla(self, datos):
        self.tabla.setRowCount(0)
        self.tabla.setUpdatesEnabled(False)
        
        for fila_data in datos:
            row_idx = self.tabla.rowCount()
            self.tabla.insertRow(row_idx)
            
            for col_idx, db_idx in enumerate(INDICES_DB_CONSULTAR):
                valor = fila_data[db_idx] if db_idx < len(fila_data) else None
                texto = str(valor) if valor is not None else ""
                self.tabla.setItem(row_idx, col_idx, QTableWidgetItem(texto))
                
        self.tabla.setUpdatesEnabled(True)

    def exportar_excel(self):
        ruta, _ = QFileDialog.getSaveFileName(
            self, "Guardar Excel", "consulta_tramites.xlsx", "Excel (*.xlsx)"
        )
        if not ruta:
            return

        wb = Workbook()
        id_memo = self.filtros.combo_memo.currentData()

        if id_memo:
            self._crear_hoja_resumen(wb)
            self._crear_hoja_documentos(wb, "Documentos")
        else:
            self._crear_hoja_documentos(wb, "Listado")

        wb.save(ruta)
        QMessageBox.information(self, "Exportado", MSG_EXCEL_EXPORTADO)

    def _crear_hoja_resumen(self, wb):
        ws = wb.active
        ws.title = "Resumen"
        ws["A1"] = MSG_TITULO_RESUMEN
        ws["A1"].font = Font(bold=True, size=14)

        if self.tabla.rowCount() == 0:
            ws["A3"] = "No hay datos cargados"
            return

        datos = [
            ("N° Trámite", self.tabla.item(0, 0).text()),
            ("Proveedor", self.tabla.item(0, 1).text()),
            ("Unidad", self.tabla.item(0, 2).text()),
            ("Fecha inicio", self.tabla.item(0, 3).text()),
        ]

        for i, (titulo, valor) in enumerate(datos, start=3):
            ws[f"A{i}"] = titulo
            ws[f"B{i}"] = valor
            ws[f"A{i}"].font = Font(bold=True)

        ws["A8"] = "Total documentos:"
        ws["B8"] = self.tabla.rowCount()

    def _crear_hoja_documentos(self, wb, nombre):
        ws = wb.create_sheet(nombre)
        ws.append(HEADER_LABELS_CONSULTAR)

        for cells in ws.iter_rows(min_row=1, max_row=1):
            for cell in cells:
                cell.font = Font(bold=True)

        for row in range(self.tabla.rowCount()):
            ws.append([
                self.tabla.item(row, col).text() 
                for col in range(self.tabla.columnCount())
            ])

        for column_cells in ws.columns:
            max_len = max(
                (len(str(cell.value or "")) for cell in column_cells),
                default=1
            )
            ws.column_dimensions[column_cells[0].column_letter].width = max_len + 2

